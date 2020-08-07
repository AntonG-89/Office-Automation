from openpyxl import Workbook, load_workbook
from mailmerge import MailMerge
import os

workbook = load_workbook(filename =r"C:\Users\anton\OneDrive\Desktop\Work Trial Build\Trial1.xlsx" )
sheet = workbook.active
max_row = sheet.max_row
max_column = sheet.max_column
template_1 = r'C:\Users\anton\OneDrive\Desktop\Work Trial Build\Trial1.docx'
template_2 = r'C:\Users\anton\OneDrive\Desktop\Work Trial Build\Trial2.docx'

def file_open():
    file_name = input("Please enter client's name and number: ?")
    







    
def template():

   
    for i in range (8,max_row+1):
        c_temp = sheet.cell(i,1)
        if c_temp.value == "CNR":
        
            document = MailMerge(template_1)
            name = sheet.cell(2,2)
            matter = sheet.cell(3,2)
            dob = sheet.cell(4,2)
            document.merge(Name = name.value)
            document.merge(Matter = str(matter.value))
            # need to add date formating
            document.merge(DOB = str(dob.value))
            for j in range (1,max_column + 1):

                cel_obj = sheet.cell(row = i,column = j)

                if j == 2:
                    document.merge(Addressee = cel_obj.value)
                    addressee = cel_obj.value

                if j == 3:
                    # need to add date formating
                    document.merge(From = str(cel_obj.value))

                if j == 4:
                    # need to add date formating
                    document.merge(to = str(cel_obj.value))
            document.write(r"C:\Users\anton\OneDrive\Desktop\CNRtemp" +addressee + ".docx")


        
        elif c_temp.value == "AB":
            
            document = MailMerge(template_2)
            name = sheet.cell(2,2)
            matter = sheet.cell(3,2)
            dob = sheet.cell(4,2)
            document.merge(Name = name.value)
            document.merge(Matter = str(matter.value))
            # need to add date formating
            document.merge(DOB = str(dob.value))
            for j in range (1,max_column + 1):

                cel_obj = sheet.cell(row = i,column = j)

                if j == 2:
                    document.merge(Addressee = cel_obj.value)
                    addressee = cel_obj.value

                if j == 3:
                    # need to add date formating
                    document.merge(From = str(cel_obj.value))

                if j == 4:
                    # need to add date formating
                    document.merge(to = str(cel_obj.value))
            document.write(r"C:\Users\anton\OneDrive\Desktop\AB" +addressee + ".docx")
           

        else:
            continue

        # Need to add another template with another value (ie. OHIP)




            
            









            
        
    


        

    
